import tweepy as tw
from textblob import TextBlob
from openpyxl import Workbook
from openpyxl.styles import Font


def paste_cells(ws_feed):
    for il in range(len(excel_pos)):
        ws_feed.cell(il*7+2, 1).value = search_list()[il]
        ws_feed.cell(il * 7 + 2, 1).font = Font(bold=True)
        ws_feed.cell(il * 7 + 2, 1).font = Font(bold=True)
        ws_feed.cell(il * 7 + 2, 1).font = Font(bold=True)
        ws_feed.cell(il*7+3, 1).value = "Positive tweets"
        ws_feed.cell(il*7+4, 1).value = "Negative tweets"
        ws_feed.cell(il*7+5, 1).value = "Neutral tweets"
        ws_feed.cell(il*7+6, 1).value = "Positive amount"
        ws_feed.cell(il*7+7, 1).value = "Negative amount"

        ws_feed.cell(il*7+3, 2).value = excel_pos[il]
        ws_feed.cell(il*7+4, 2).value = excel_neg[il]
        ws_feed.cell(il*7+5, 2).value = excel_neu[il]
        ws_feed.cell(il*7+6, 2).value = excel_pos_amt[il]
        ws_feed.cell(il*7+7, 2).value = excel_neg_amt[il]


def paste_tweets(ws_tweets):
    for iu in range(len(excel_all_tweets)):
        ws_tweets.cell(1, iu*3+1).value = search_list()[iu]
        ws_tweets.cell(1, iu*3+1).font = Font(bold=True)
        ws_tweets.cell(1, iu*3+2).value = "Date"
        ws_tweets.cell(1, iu*3+2).font = Font(bold=True)
        for iy in range(tweet_amount):
            try:
                ws_tweets.cell(iy+2, iu*3+1).value = excel_all_tweets[iu][iy]
                ws_tweets.cell(iy+2, iu*3+2).value = excel_date[iu][iy]
            except:
                pass

def paste_excel():
    ws = wb.active
    ws.title = "Twitter feed"
    ws_feed = wb["Twitter feed"]
    ws_tweets = wb.create_sheet("All_tweets")
    paste_tweets(ws_tweets)
    paste_cells(ws_feed)


def search_list():
    temp_list = []
    user_hashtags = b_user_hashtags.split(",")
    for i in range(len(user_hashtags)):
        temp_list.append("#" + user_hashtags[i])
    return temp_list


def twitter_search():
    x = 2


consumer_key = "Use your key here"
consumer_secret = "Use your key here"
access_token = "Use your key here"
access_token_secret = "Use your key here"

auth = tw.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(access_token, access_token_secret)
api = tw.API(auth, wait_on_rate_limit=True)

user_save = input("Enter project name: ")
loc_save = input(r"Enter where to save the project: [C:\Users\..]")
backslash_sol = "\\"
print(f"Project will be saved as {loc_save + backslash_sol + user_save}.xlsx")
wb = Workbook()

b_user_hashtags = input("Enter the hashtags you would like to search for [separated by , and no spaces]: ")
date_since = "2015-09-13"
print_list = []


print(search_list())
excel_neg, excel_neu, excel_pos, excel_neg_amt, excel_pos_amt, excel_all_tweets, excel_date = [], [], [], [], [], [], []
tweet_amount = int(input("How many tweets would you like to search: "))

for hashes in range(len(search_list())):
    #new_search = search_list() + " -filter:retweets"
    tweet_list, tweet_location, sentiment_list = [], [], []
    excel_all_tweets.append([])
    excel_date.append([])
    tweet_count = 0
    neg_count, pos_count, neu_count = 0, 0, 0
    neg_amt, pos_amt, neu_amt = 0, 0, 0
    # Collect tweets
    tweets = tw.Cursor(api.search, q=search_list()[hashes] + "-filter:retweets", lang="en", since=date_since).items(tweet_amount)
    for tweet in tweets:
        try:
            tweet_list.append(tweet.text)
            tweet_location.append(tweet.user.location)
            print(f"[{tweet_count+1}/{tweet_amount}] {tweet_location[tweet_count]} \n {tweet.text}")
            tweet_count += 1
            analysis = TextBlob(tweet.text)
            print(analysis.sentiment)
            sentiment_list.append(analysis.sentiment)
            excel_all_tweets[hashes].append(tweet.text)
            excel_date[hashes].append(tweet.created_at)
            if analysis.sentiment[0] == 0:
                neu_amt += analysis.sentiment[1]
                neu_count += 1
            elif analysis.sentiment[0] > 0:
                pos_count += 1
                pos_amt += analysis.sentiment[1]
            else:
                neg_count += 1
                neg_amt += analysis.sentiment[1]
        except:
            break
    print_list.append(f"\nFor {search_list()[hashes]}: \nNegative tweets count: {neg_count}\nPositive tweets count: {pos_count}\nNeutral tweets count: {neu_count}\nNegative tweets amount: {neg_amt}\nPositive tweets amount: {pos_amt}")
    excel_neg.append(neg_count)
    excel_neu.append(neu_count)
    excel_pos.append(pos_count)
    excel_neg_amt.append(neg_amt)
    excel_pos_amt.append(pos_amt)
    wb.save(loc_save[:-1] + user_save + ".xlsx")

for number_hashes in range(len(print_list)):
    print(print_list[number_hashes])

paste_excel()
wb.save(loc_save[:-1] + user_save + ".xlsx")

print(f"Project saved as {loc_save + backslash_sol + user_save}.xlsx")




